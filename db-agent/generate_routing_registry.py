"""
DB Agent — Routing Registry Generator
======================================
Parses the Oracle PL/SQL package definitions embedded in REGISTRY_SOURCE below
to extract exact JSON key names, then writes a formatted Excel routing registry.

Run from db-agent/:
    python generate_routing_registry.py

Output:
    db-agent/routing_registry.xlsx
"""

import datetime
import json
import re

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# ── OUTPUT PATH ──────────────────────────────────────────────────────────────

OUTPUT_FILE = "routing_registry.xlsx"


# ── PL/SQL SOURCE — PARSED TO EXTRACT JSON KEYS ──────────────────────────────
# Paste PL/SQL package bodies here; the extractor reads JSON_VALUE key paths
# (format: '$.key_name') and JSON_OBJECT key names (format: 'key' VALUE ...).

REGISTRY_SOURCE = {
    # -------------------------------------------------------------------------
    # Source PL/SQL for the first registry row — provided verbatim by the user.
    # The extractor reads every JSON_VALUE call's key path ($.claim_id, etc.)
    # as an input field, and every JSON_OBJECT 'key' VALUE token as an output.
    # -------------------------------------------------------------------------
    "PKG_AGENT_INSERT_CLAIM.PROCESS_ATTACHMENT_JSON": """
        CREATE OR REPLACE PACKAGE BODY PKG_AGENT_INSERT_CLAIM AS
            PROCEDURE PROCESS_ATTACHMENT_JSON (
                p_json_payload   IN  CLOB,
                p_json_response  OUT CLOB
            ) IS
                v_claim_id         VARCHAR2(255);
                v_attachment_name  VARCHAR2(255);
                v_attachment_order NUMBER;
                v_version          NUMBER;
                v_page_count       NUMBER;
                v_extracted_text   CLOB;
                v_row_id           NUMBER;
                v_error_msg        VARCHAR2(4000);
            BEGIN
                v_claim_id         := JSON_VALUE(p_json_payload, '$.claim_id');
                v_attachment_name  := JSON_VALUE(p_json_payload, '$.attachment_name');
                v_attachment_order := TO_NUMBER(JSON_VALUE(p_json_payload, '$.attachment_order'));
                v_version          := TO_NUMBER(JSON_VALUE(p_json_payload, '$.version'));
                v_page_count       := TO_NUMBER(JSON_VALUE(p_json_payload, '$.page_count'));
                v_extracted_text   := JSON_VALUE(p_json_payload, '$.extracted_text' RETURNING CLOB);

                PRC_CM_UPS_CLAIM_ATTACHMENT(
                    p_claim_id         => v_claim_id,
                    p_attachment_name  => v_attachment_name,
                    p_attachment_order => v_attachment_order,
                    p_version          => v_version,
                    p_page_count       => v_page_count,
                    p_extracted_text   => v_extracted_text,
                    p_row_id           => v_row_id
                );

                SELECT JSON_OBJECT(
                    'status'  VALUE 'SUCCESS',
                    'message' VALUE 'Attachment processed successfully.',
                    'row_id'  VALUE v_row_id
                ) INTO p_json_response FROM DUAL;
            END PROCESS_ATTACHMENT_JSON;
        END PKG_AGENT_INSERT_CLAIM;
    """,
}


# ── PL/SQL KEY EXTRACTOR ─────────────────────────────────────────────────────

def extract_plsql_keys(plsql: str):
    """
    Scans PL/SQL source text for two patterns and returns the extracted key names.

    Input keys:  JSON_VALUE(p_json_payload, '$.key_name') calls
                 → these become the expected IN CLOB JSON payload fields
    Output keys: JSON_OBJECT('key' VALUE ...) expressions in SELECT INTO
                 → these become the expected OUT CLOB JSON response fields

    Returns:
        (input_keys: list[str], output_keys: list[str])
    """

    # pattern: JSON_VALUE( anything, '$.KEY_NAME' optional-stuff )
    # captures the key name after the $. prefix
    input_pattern  = re.compile(r"JSON_VALUE\s*\([^,]+,\s*'\$\.(\w+)", re.IGNORECASE)

    # pattern: 'KEY_NAME' VALUE  — used in JSON_OBJECT('key' VALUE expr, ...)
    # only captures keys that immediately precede the VALUE keyword
    output_pattern = re.compile(r"'(\w+)'\s+VALUE", re.IGNORECASE)

    # deduplicate while preserving order (dict trick — Python 3.7+)
    input_keys  = list(dict.fromkeys(input_pattern.findall(plsql)))
    output_keys = list(dict.fromkeys(output_pattern.findall(plsql)))

    return input_keys, output_keys


# ── REGISTRY ROWS ─────────────────────────────────────────────────────────────
# Row 1 is built by parsing the PL/SQL above.
# Rows 2-4 are representative mock entries that show how the registry scales.

def build_rows():
    """Constructs the list of row dicts that will populate the Excel sheet."""

    rows = []

    # ── ROW 1: Oracle PL/SQL procedure — parsed from REGISTRY_SOURCE ─────────

    proc_key = "PKG_AGENT_INSERT_CLAIM.PROCESS_ATTACHMENT_JSON"
    plsql_src = REGISTRY_SOURCE[proc_key]

    # extract the exact JSON field names from the PL/SQL source
    in_keys, out_keys = extract_plsql_keys(plsql_src)

    # build a realistic example payload using the extracted key names
    # values are representative samples, not real data
    sample_payload = {k: _sample_value(k) for k in in_keys}

    # build the expected response shape using the extracted output key names
    # values match what the JSON_OBJECT 'VALUE' literal expressions produce
    sample_response_data = {
        "status":  "SUCCESS",
        "message": "Attachment processed successfully.",
        "row_id":  98231,
    }

    # the /execute-procedure endpoint wraps the body inside a { status, data } envelope
    api_payload = {
        "target_procedure": proc_key,
        "token":            "<INTERNAL_SERVICE_TOKEN>",
        "service_name":     "claims-expert",
        "payload":          sample_payload,
    }

    api_response = {"status": "SUCCESS", "data": sample_response_data}

    rows.append({
        "Route Name":             f"pkg_procedure:{proc_key}",
        "API Endpoint":           "/execute-procedure",
        "HTTP Method":            "POST",
        "Target DB":              "Oracle 23ai — VECTOR_DB\n192.168.244.198:1528",
        "DB Type":                "Oracle 23ai",
        "Package":                "PKG_AGENT_INSERT_CLAIM",
        "Procedure":              "PROCESS_ATTACHMENT_JSON",
        "Parsed Input Keys":      ", ".join(in_keys),
        "Parsed Output Keys":     ", ".join(out_keys),
        "Input JSON Payload":     json.dumps(api_payload, indent=2),
        "Expected Response JSON": json.dumps(api_response, indent=2),
        "Description": (
            "Upserts a claim attachment record in Oracle 23ai. "
            "Parses the CLOB JSON payload, calls PRC_CM_UPS_CLAIM_ATTACHMENT, "
            "and returns the generated row_id. "
            f"Input keys extracted from PL/SQL: {', '.join(in_keys)}."
        ),
    })

    # ── ROW 2: Oracle 23ai — legacy static procedure route ───────────────────
    # 'process_attachment' is the original static route in ROUTE_REGISTRY that
    # calls PKG_AGENT_INSERT_CLAIM via oracle_db/oracle_connector.py (not vector_connector).
    # Kept alongside the newer pkg_procedure: generic route for backwards compatibility.

    rows.append({
        "Route Name":             "process_attachment",
        "API Endpoint":           "/claims/store",
        "HTTP Method":            "POST",
        "Target DB":              "Oracle 23ai — VECTOR_DB\n192.168.244.198:1528",
        "DB Type":                "Oracle 23ai",
        "Package":                "PKG_AGENT_INSERT_CLAIM",
        "Procedure":              "PROCESS_ATTACHMENT_JSON",
        "Parsed Input Keys":      "claim_id, attachment_name, attachment_order, version, page_count, extracted_text",
        "Parsed Output Keys":     "status, message, row_id",
        "Input JSON Payload": json.dumps({
            "claim_id":         "CLM-2025-001234",
            "attachment_name":  "EOB_page1.pdf",
            "attachment_order": 1,
            "version":          2,
            "page_count":       4,
            "extracted_text":   "Patient: John Doe. DOS: 2025-01-15...",
        }, indent=2),
        "Expected Response JSON": json.dumps({
            "status": "SUCCESS",
            "data": {
                "status":  "SUCCESS",
                "message": "Attachment processed successfully.",
                "row_id":  98231,
            },
        }, indent=2),
        "Description": (
            "Static route — calls PKG_AGENT_INSERT_CLAIM.PROCESS_ATTACHMENT_JSON "
            "via oracle_db/oracle_connector.py (not the generic /execute-procedure path). "
            "Used by /claims/store. Kept for backwards compatibility."
        ),
    })

    # ── ROW 3: NL2SQL — translate + execute on LEGACY, DWH, or VECTOR ─────────
    # Route name must be exactly 'nl_query' — that is the key in ROUTE_REGISTRY.
    # target_db=VECTOR triggers vector search; LEGACY/DWH generate SQL via LLM Agent.

    rows.append({
        "Route Name":             "nl_query",
        "API Endpoint":           "/query/nl",
        "HTTP Method":            "POST",
        "Target DB":              "LEGACY (PostgreSQL) | DWH (PostgreSQL) | VECTOR (Oracle 23ai)",
        "DB Type":                "PostgreSQL / Oracle 23ai",
        "Package":                "N/A — LLM Agent NL2SQL + asyncpg / VECTOR_DISTANCE",
        "Procedure":              "N/A — dynamic SQL or ANN search",
        "Parsed Input Keys":      "query, target_db, query_vector, top_k",
        "Parsed Output Keys":     "target_db, sql, results  |  target_db, results[ doc_id, score, payload ]",
        "Input JSON Payload": json.dumps({
            "query":        "Show all approved claims from last month",
            "target_db":    "LEGACY",
            "query_vector": None,
            "top_k":        5,
        }, indent=2),
        "Expected Response JSON": json.dumps({
            "status": "SUCCESS",
            "data": {
                "target_db": "LEGACY",
                "sql":       "SELECT * FROM claims WHERE status='APPROVED' AND ...",
                "results":   [
                    {"claim_id": "CLM-2025-001100", "status": "APPROVED", "amount": 870.00},
                ],
            },
        }, indent=2),
        "Description": (
            "Translates a natural language question to SQL via LLM Agent (port 8005), "
            "validates it with is_safe_query() (SELECT-only whitelist), "
            "then executes it on the requested target_db. "
            "If target_db=VECTOR, skips SQL generation and performs cosine ANN search instead."
        ),
    })

    # ── ROW 4: Legacy PostgreSQL — member claim history ───────────────────────

    rows.append({
        "Route Name":             "fetch_member_history",
        "API Endpoint":           "/claims/history",
        "HTTP Method":            "POST",
        "Target DB":              "Legacy PostgreSQL — LEGACY_DB",
        "DB Type":                "PostgreSQL",
        "Package":                "N/A",
        "Procedure":              "N/A — asyncpg SELECT",
        "Parsed Input Keys":      "member_id",
        "Parsed Output Keys":     "member_id, count, claims[ claim_id, status, amount ]",
        "Input JSON Payload": json.dumps({
            "member_id": "MBR-100293",
        }, indent=2),
        "Expected Response JSON": json.dumps({
            "status": "SUCCESS",
            "data": {
                "member_id": "MBR-100293",
                "count": 3,
                "claims": [
                    {"claim_id": "CLM-2024-009901", "status": "APPROVED", "amount": 1250.00},
                    {"claim_id": "CLM-2025-000011", "status": "PENDING",  "amount":  340.00},
                    {"claim_id": "CLM-2025-001100", "status": "DENIED",   "amount":  870.00},
                ],
            },
        }, indent=2),
        "Description": (
            "Fetches the full claim history for a given member "
            "from the legacy on-premise PostgreSQL database. "
            "Results are ordered by claim date descending."
        ),
    })

    # ── ROW 5: Legacy PostgreSQL — authorization lookup ───────────────────────

    rows.append({
        "Route Name":             "fetch_authorization",
        "API Endpoint":           "/authorization/fetch",
        "HTTP Method":            "POST",
        "Target DB":              "Legacy PostgreSQL — LEGACY_DB",
        "DB Type":                "PostgreSQL",
        "Package":                "N/A",
        "Procedure":              "N/A — asyncpg SELECT",
        "Parsed Input Keys":      "claim_id",
        "Parsed Output Keys":     "found, claim_id, authorized, coverage_type, max_benefit, used_benefit",
        "Input JSON Payload": json.dumps({
            "claim_id": "CLM-2025-001234",
        }, indent=2),
        "Expected Response JSON": json.dumps({
            "status": "SUCCESS",
            "data": {
                "found":         True,
                "claim_id":      "CLM-2025-001234",
                "authorized":    True,
                "coverage_type": "IN_NETWORK",
                "max_benefit":   5000.00,
                "used_benefit":  1250.00,
            },
        }, indent=2),
        "Description": (
            "Retrieves authorization and insurance coverage data for a specific claim. "
            "Returns found=false if no authorization record exists. "
            "Used by the Orchestrator Flight Plan to gate claim processing."
        ),
    })

    # ── ROW 6: Logging PostgreSQL — audit event write ─────────────────────────

    rows.append({
        "Route Name":             "write_audit_log",
        "API Endpoint":           "/logs/audit",
        "HTTP Method":            "POST",
        "Target DB":              "Logging PostgreSQL — LOGGING_DB",
        "DB Type":                "PostgreSQL",
        "Package":                "N/A",
        "Procedure":              "N/A — asyncpg INSERT",
        "Parsed Input Keys":      "event_type, service_name, claim_id, details",
        "Parsed Output Keys":     "written",
        "Input JSON Payload": json.dumps({
            "event_type":   "CLAIM_PROCESSED",
            "service_name": "claims-expert",
            "claim_id":     "CLM-2025-001234",
            "details":      {"step": "ocr_complete", "pages": 4},
        }, indent=2),
        "Expected Response JSON": json.dumps({
            "status": "SUCCESS",
            "data":   {"written": True},
        }, indent=2),
        "Description": (
            "Appends one structured audit event to the logging PostgreSQL database. "
            "Called after every significant Flight Plan step so operations can "
            "reconstruct the full claim processing timeline."
        ),
    })

    # ── ROW 7: Logging PostgreSQL — save Orchestrator service registry ─────────

    rows.append({
        "Route Name":             "save_registry",
        "API Endpoint":           "/registry/save",
        "HTTP Method":            "POST",
        "Target DB":              "Logging PostgreSQL — LOGGING_DB",
        "DB Type":                "PostgreSQL",
        "Package":                "N/A",
        "Procedure":              "N/A — asyncpg INSERT (JSONB)",
        "Parsed Input Keys":      "registry{ service_name: url }",
        "Parsed Output Keys":     "saved, service_count",
        "Input JSON Payload": json.dumps({
            "registry": {
                "claims-expert":   "http://claims-expert:8004",
                "llm-agent":       "http://llm-agent:8005",
                "ocr-agent":       "http://ocr-agent:8006",
            },
        }, indent=2),
        "Expected Response JSON": json.dumps({
            "status": "SUCCESS",
            "data":   {"saved": True, "service_count": 3},
        }, indent=2),
        "Description": (
            "Persists a snapshot of the Orchestrator's live service registry "
            "to the logging database as JSONB. "
            "Called by the Orchestrator after every service registration change "
            "so the registry can be restored after a restart."
        ),
    })

    # ── ROW 8: Logging PostgreSQL — load Orchestrator service registry ─────────

    rows.append({
        "Route Name":             "load_registry",
        "API Endpoint":           "/registry/load",
        "HTTP Method":            "GET",
        "Target DB":              "Logging PostgreSQL — LOGGING_DB",
        "DB Type":                "PostgreSQL",
        "Package":                "N/A",
        "Procedure":              "N/A — asyncpg SELECT (JSONB)",
        "Parsed Input Keys":      "(none — GET request, no body)",
        "Parsed Output Keys":     "registry{ service_name: url }, service_count",
        "Input JSON Payload":     "(no body)",
        "Expected Response JSON": json.dumps({
            "status": "SUCCESS",
            "data": {
                "registry": {
                    "claims-expert": "http://claims-expert:8004",
                    "llm-agent":     "http://llm-agent:8005",
                },
                "service_count": 2,
            },
        }, indent=2),
        "Description": (
            "Loads the most recently saved service registry snapshot from the "
            "logging database. Called by the Orchestrator on startup. "
            "Returns an empty registry dict if no snapshot has been saved yet."
        ),
    })

    return rows


def _sample_value(key: str):
    """
    Returns a realistic sample value for a given JSON key name.
    Used when constructing the example payload from PL/SQL-parsed field names.
    """
    # map known key patterns to representative sample values
    samples = {
        "claim_id":         "CLM-2025-001234",
        "attachment_name":  "EOB_page1.pdf",
        "attachment_order": 1,
        "version":          2,
        "page_count":       4,
        "extracted_text":   "Patient: John Doe. Date of service: 2025-01-15. "
                            "Procedure: 99213 Office visit...",
    }
    # fall back to a generic string if the key is not in the lookup
    return samples.get(key, f"<{key}>")


# ── EXCEL BUILDER ─────────────────────────────────────────────────────────────

# defines the columns that appear in the spreadsheet — order matters
COLUMNS = [
    "Route Name",
    "API Endpoint",
    "HTTP Method",
    "Target DB",
    "DB Type",
    "Package",
    "Procedure",
    "Parsed Input Keys",
    "Parsed Output Keys",
    "Input JSON Payload",
    "Expected Response JSON",
    "Description",
]

# column widths in Excel character units
COL_WIDTHS = {
    "Route Name":             52,
    "API Endpoint":           22,
    "HTTP Method":            13,
    "Target DB":              28,
    "DB Type":                14,
    "Package":                34,
    "Procedure":              30,
    "Parsed Input Keys":      36,
    "Parsed Output Keys":     42,
    "Input JSON Payload":     54,
    "Expected Response JSON": 54,
    "Description":            52,
}

# columns rendered in Courier New for machine-readable content
MONOSPACE_COLS = {
    "Route Name", "Package", "Procedure",
    "Parsed Input Keys", "Parsed Output Keys",
    "Input JSON Payload", "Expected Response JSON",
}


def _thin_border():
    """Returns a uniform thin-grey border for all data cells."""
    side = Side(style="thin", color="CCCCCC")
    return Border(left=side, right=side, top=side, bottom=side)


def build_excel(rows):
    """Writes the registry rows to a formatted Excel workbook."""

    wb  = openpyxl.Workbook()
    ws  = wb.active
    ws.title = "Routing Registry"

    border  = _thin_border()
    n_cols  = len(COLUMNS)

    # ── ROW 1: WORKBOOK TITLE ─────────────────────────────────────────────────

    title_cell = ws.cell(row=1, column=1, value="DB Agent — Routing Registry")
    title_cell.font      = Font(name="Calibri", bold=True, size=16, color="1F4E79")
    title_cell.fill      = PatternFill("solid", fgColor="D6E4F0")
    title_cell.alignment = Alignment(horizontal="left", vertical="center",
                                     indent=1)
    # merge the title across all columns so it spans the full width
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    ws.row_dimensions[1].height = 36

    # ── ROW 2: METADATA SUBTITLE ──────────────────────────────────────────────

    subtitle = (
        f"Generated: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M')}  |  "
        f"Service: db-agent  |  Port: 8003  |  "
        f"Oracle schema: mldsdev  |  DB: aidev_pdb1"
    )
    sub_cell = ws.cell(row=2, column=1, value=subtitle)
    sub_cell.font      = Font(name="Calibri", italic=True, size=10, color="595959")
    sub_cell.fill      = PatternFill("solid", fgColor="EBF5FB")
    sub_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n_cols)
    ws.row_dimensions[2].height = 20

    # ── ROW 3: COLUMN HEADERS ─────────────────────────────────────────────────

    hdr_font  = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    hdr_fill  = PatternFill("solid", fgColor="1F4E79")
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col_idx, col_name in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=3, column=col_idx, value=col_name)
        cell.font      = hdr_font
        cell.fill      = hdr_fill
        cell.alignment = hdr_align
        cell.border    = border

    ws.row_dimensions[3].height = 32

    # ── ROWS 4+: DATA ─────────────────────────────────────────────────────────

    plain_font = Font(name="Calibri", size=10)
    mono_font  = Font(name="Courier New", size=9)
    wrap_top   = Alignment(vertical="top", wrap_text=True)

    # alternate row shading for readability
    fills = [
        PatternFill("solid", fgColor="EAF2FB"),   # even rows — light blue
        PatternFill("solid", fgColor="FFFFFF"),    # odd rows  — white
    ]

    for row_idx, row_data in enumerate(rows, start=4):
        fill = fills[(row_idx) % 2]   # alternates every row

        for col_idx, col_name in enumerate(COLUMNS, start=1):
            value = row_data.get(col_name, "")
            cell  = ws.cell(row=row_idx, column=col_idx, value=value)

            cell.fill      = fill
            cell.border    = border
            cell.alignment = wrap_top
            # monospace for code/JSON columns, proportional for prose
            cell.font = mono_font if col_name in MONOSPACE_COLS else plain_font

        # tall rows to accommodate multi-line JSON blocks
        ws.row_dimensions[row_idx].height = 155

    # ── COLUMN WIDTHS ─────────────────────────────────────────────────────────

    for col_idx, col_name in enumerate(COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = COL_WIDTHS[col_name]

    # ── FREEZE PANES & ZOOM ───────────────────────────────────────────────────

    # freezes the title, subtitle, and header rows so they remain visible while scrolling
    ws.freeze_panes   = "A4"
    # slightly zoomed out so the wide JSON columns stay visible without horizontal scrolling
    ws.sheet_view.zoomScale = 85

    # ── SAVE ──────────────────────────────────────────────────────────────────

    wb.save(OUTPUT_FILE)

    print(f"Registry written → {OUTPUT_FILE}")
    print(f"  Rows  : {len(rows)}")
    print(f"  Cols  : {n_cols}")
    print()
    for row in rows:
        in_keys  = row.get("Parsed Input Keys",  "—")
        out_keys = row.get("Parsed Output Keys", "—")
        print(f"  [{row['Route Name'][:55]}]")
        print(f"    IN  keys : {in_keys}")
        print(f"    OUT keys : {out_keys}")
        print()


# ── ENTRY POINT ───────────────────────────────────────────────────────────────

if __name__ == "__main__":
    rows = build_rows()
    build_excel(rows)
